import query.connect as connect
import query.constants as constants
from pathlib import Path
from collections.abc import Iterable, Mapping
from openpyxl.styles import Alignment
import openpyxl

def tables_generate(
    students: Iterable[tuple[int, str]],
    lessons: Iterable[dict],
    marks: Mapping[tuple[int, int], dict],
    info: dict,
    save_path: str,
):
    template_path = Path("export/samples/tables.xlsx")
    wb = openpyxl.load_workbook(template_path)
    sheet = wb["Sheet1"]

    # Шапка документа
    sheet["A2"] = connect.config["university"]["status"]
    sheet["A3"] = connect.config["university"]["full_name"]
    sheet["A4"] = connect.config["university"]["institute"]
    sheet["A5"] = connect.config["university"]["departament"]

    sheet["A7"] = "Дисциплина: " + info['name'] + " (" + constants.LESSON_TYPE_SMALL[info['type']] + ")"
    sheet["A8"] = "Часы: " + str(info['hours'])
    sheet['A9'] = "Преподаватель: " + info['teacher']

    HEADER_ROW = 11
    DATE_TITLE_ROW = 10
    FIRST_STUDENT_ROW = 12
    FIRST_LESSON_COLUMN = 2

    students = list(students)
    lessons = list(lessons)

    lesson_count = len(lessons)

    if lesson_count > 0:
        last_lesson_column = FIRST_LESSON_COLUMN + lesson_count - 1

        sheet.merge_cells(
            start_row=DATE_TITLE_ROW,
            start_column=FIRST_LESSON_COLUMN,
            end_row=DATE_TITLE_ROW,
            end_column=last_lesson_column,
        )

        date_title_cell = sheet.cell(
            row=DATE_TITLE_ROW,
            column=FIRST_LESSON_COLUMN,
            value="Дата",
        )

        date_title_cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
        )

    # Студенты располагаются подряд, независимо от ID в БД
    for row_offset, (student_id, student_name) in enumerate(students):
        row = FIRST_STUDENT_ROW + row_offset
        sheet.cell(row=row, column=1, value=student_name)

    # Занятия располагаются подряд, независимо от ID в БД
    for column_offset, lesson in enumerate(lessons):
        column = FIRST_LESSON_COLUMN + column_offset
        sheet.cell(
            row=HEADER_ROW,
            column=column,
            value=lesson["date"].strftime("%d.%m"),
        )

    # Заполнение посещаемости
    for row_offset, (student_id, _) in enumerate(students):
        row = FIRST_STUDENT_ROW + row_offset

        for column_offset, lesson in enumerate(lessons):
            column = FIRST_LESSON_COLUMN + column_offset
            lesson_id = lesson["id"]

            # get() не вызывает KeyError, если отметки нет
            mark_info = marks.get((student_id, lesson_id))
            mark_value = mark_info["mark"] if mark_info else ""

            sheet.cell(
                row=row,
                column=column,
                value=mark_value,
            )

    wb.save(save_path)